import streamlit as st
import pdfplumber
import re
import pandas as pd
from io import BytesIO
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.set_page_config(page_title="Leitor Cartão de Ponto", layout="centered")
st.title("📄 Leitor Inteligente - Cartão de Ponto")
st.write("Envie o PDF (TPBR ou JPBB).")

uploaded_file = st.file_uploader("Enviar PDF", type=["pdf"])


# ==========================
# Funções auxiliares
# ==========================
def hhmm_to_minutes(hhmm: str) -> int:
    if not hhmm or ":" not in hhmm:
        return 0
    h, m = hhmm.split(":")
    return int(h) * 60 + int(m)


def minutes_to_hhmm(minutes: int) -> str:
    sign = "-" if minutes < 0 else ""
    minutes = abs(minutes)
    return f"{sign}{minutes // 60:02d}:{minutes % 60:02d}"


# ==========================
# Leitura com CACHE
# ==========================
@st.cache_data
def extract_full_text(pdf_file) -> str:
    with pdfplumber.open(pdf_file) as pdf:
        parts = []
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
    return "\n".join(parts)


@st.cache_data
def parse_employee_blocks(texto: str):
    blocos = re.split(r"\bCartão\s+de\s+Ponto\b", texto)
    out = []

    for bloco in blocos:
        if "NOME DO FUNCIONÁRIO:" not in bloco or "TOTAIS" not in bloco:
            continue

        nome_match = re.search(r"NOME DO FUNCIONÁRIO:\s*(.+?)\s+PIS", bloco)
        if not nome_match:
            continue
        nome = nome_match.group(1).strip()

        cargo_match = re.search(r"NOME DO CARGO:\s*(.+)", bloco)
        cargo = cargo_match.group(1).split("\n")[0].strip().upper() if cargo_match else ""

        totais_match = re.search(r"TOTAIS\s+([0-9:\s]+)", bloco)
        if not totais_match:
            continue

        horarios = re.findall(r"\d{1,3}:\d{2}", totais_match.group(1))

        total_normais = "00:00"
        total_noturno = "00:00"
        falta = "00:00"
        extra70 = "00:00"

        if len(horarios) == 5:
            _, total_normais, total_noturno, falta, extra70 = horarios
        elif len(horarios) == 4:
            total_normais, total_noturno, falta, extra70 = horarios
        elif len(horarios) >= 6:
            _, total_normais, total_noturno, falta, extra70 = horarios[:5]
        elif len(horarios) == 3:
            total_normais, total_noturno, extra70 = horarios
        elif len(horarios) == 2:
            total_normais, extra70 = horarios
        elif len(horarios) == 1:
            total_normais = horarios[0]

        out.append({
            "NOME": nome,
            "CARGO": cargo,
            "TOTAL NORMAIS": total_normais,
            "TOTAL NOTURNO": total_noturno,
            "FALTA": falta,
            "EXTRA 70%": extra70,
        })

    return out


# ==========================
# Excel Formatado
# ==========================
def build_excel(df_func, df_moto):
    wb = Workbook()
    ws = wb.active
    ws.title = "RELATORIO"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_gray = PatternFill(start_color="BFBFBF", fill_type="solid")
    title_yellow = PatternFill(start_color="FFFF00", fill_type="solid")

    def style_cell(cell, bold=False, fill=None):
        cell.font = Font(bold=bold)
        cell.alignment = Alignment(horizontal="center")
        if fill:
            cell.fill = fill
        cell.border = border

    start_row = 2

    for col, name in enumerate(df_func.columns, start=1):
        c = ws.cell(row=start_row, column=col, value=name)
        style_cell(c, bold=True, fill=header_gray)

    for r_idx, row in enumerate(df_func.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=col_idx, value=value)
            style_cell(c)

    row_title = start_row + len(df_func) + 3
    ws.merge_cells(start_row=row_title, start_column=1, end_row=row_title, end_column=4)
    t = ws.cell(row=row_title, column=1, value="MOTOBOYS HORISTAS")
    style_cell(t, bold=True, fill=title_yellow)
    ws.row_dimensions[row_title].height = 20

    header_row2 = row_title + 1

    for col, name in enumerate(df_moto.columns, start=1):
        c = ws.cell(row=header_row2, column=col, value=name)
        style_cell(c, bold=True, fill=header_gray)

    for r_idx, row in enumerate(df_moto.itertuples(index=False), start=header_row2 + 1):
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=col_idx, value=value)
            style_cell(c)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ==========================
# PROCESSAMENTO
# ==========================
if uploaded_file and st.button("🚀 Processar PDF"):

    progress = st.progress(0)
    status = st.empty()

    status.text("🔄 Lendo PDF...")
    texto = extract_full_text(uploaded_file)
    time.sleep(0.3)
    progress.progress(40)

    status.text("📊 Processando funcionários...")
    dados = parse_employee_blocks(texto)
    time.sleep(0.3)
    progress.progress(80)

    if not dados:
        st.error("Não encontrei funcionários no PDF.")
        st.stop()

    df = pd.DataFrame(dados)

    df["EXTRA OU FALTA"] = (
        df["EXTRA 70%"].apply(hhmm_to_minutes)
        - df["FALTA"].apply(hhmm_to_minutes)
    ).apply(minutes_to_hhmm)

    is_motoboy = df["CARGO"].str.contains("MOTOBOY", na=False)

    df_func = df[~is_motoboy][["NOME", "FALTA", "EXTRA 70%", "EXTRA OU FALTA", "TOTAL NOTURNO"]].copy()
    df_func.columns = ["NOME", "FALTA", "EXTRA", "EXTRA OU FALTA", "NOTURNO"]

    df_moto_raw = df[is_motoboy].copy()
    df_moto = df_moto_raw[["NOME", "TOTAL NOTURNO", "TOTAL NORMAIS", "EXTRA 70%"]].copy()
    df_moto.columns = ["NOME", "NOTURNO", "HORAS", "EXTRA"]

    status.text("✅ Finalizado!")
    progress.progress(100)

    st.divider()

    st.subheader("📊 RESUMO GERAL")

    col1, col2 = st.columns(2)
    col1.metric("Funcionários", len(df_func))
    col2.metric("Motoboys", len(df_moto))

    total_extra = df_func["EXTRA"].apply(hhmm_to_minutes).sum()
    total_falta = df_func["FALTA"].apply(hhmm_to_minutes).sum()

    col3, col4 = st.columns(2)
    col3.metric("Total Extra", minutes_to_hhmm(total_extra))
    col4.metric("Total Falta", minutes_to_hhmm(total_falta))

    st.divider()

    st.subheader("FUNCIONÁRIOS")
    st.dataframe(df_func, use_container_width=True)

    st.subheader("MOTOBOYS HORISTAS")
    st.dataframe(df_moto, use_container_width=True)

    excel_buffer = build_excel(df_func, df_moto)

    st.download_button(
        "⬇️ Baixar Excel",
        data=excel_buffer,
        file_name="Relatorio_Modelo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
